import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import xlrd
from datetime import datetime
import time
import re
import glob
import os
from openpyxl import workbook
from openpyxl import load_workbook


def locationTaskReminder():
    # todo latest TaskReminder File in the same folder
    # open the latest TaskReminder file in the parent folder
    list_of_files = glob.glob('CRQ\T*.xlsx')
    latest_file = max(list_of_files, key=os.path.getctime)
    loc = latest_file  # Name and location of the file with names of the assignees
    wb = xlrd.open_workbook(loc)  # Open wb as excel sheet
    sheet = wb.sheet_by_index(0)  # starts reading from origin

    with open("output.txt") as openfile:
        for line in openfile:
            if loc in line:
                exit()

    with open("output.txt", "w") as text_file:
        print(loc, file=text_file)

    return sheet


def locationDCNEmailAddress():
    # todo DCN Team Email address Workbook assignment
    # open the Email addresses file in the parent folder
    locAddr = "DCN Team Email address.xlsx"  # Name and location of the file which contains email addresses
    wbAddr = xlrd.open_workbook(locAddr)  # Open wbAddr as excel sheet
    sheetAddr = wbAddr.sheet_by_index(0)  # starts reading from origin
    return sheetAddr


def locationMasterSpreadSheet_Read():
    # todo open MasterSpreadSheet for reading purposes
    loc_write = "MasterSpreadSheet.xlsx"
    wb_write = xlrd.open_workbook(loc_write)
    sheet_write = wb_write.sheet_by_index(0)
    return sheet_write


def locationMasterSpreadSheet_Write():
    # todo open MasterSpreadSheet for writing purposes
    wb = load_workbook("MasterSpreadSheet.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    return Sheet1


def Email_System(sheet, sheetAddr):
    # todo This function is responsible for sending emails based on some conditions by reading the content from the
    # todo taskreminder file

    for count in range(sheet.nrows - 1):
        loopBreak = 0
        if sheet.cell_value(count + 1, 0) != '':
            names = (sheet.cell_value(count + 1, 5))  # name of the assignee
            cid = (sheet.cell_value(count + 1, 0))  # Change ID
            tid = (sheet.cell_value(count + 1, 1))  # task ID
            time_value = float(sheet.cell_value(count + 1, 6))  # Time is in float
            x = time_value
            datetime_value = datetime(*xlrd.xldate_as_tuple(x, 0))  # Converts to date time format
            timeNow = f"{datetime.now():%Y-%d-%m %H:%M:%S}"
            datetime_object = datetime.strptime(timeNow, '%Y-%d-%m %H:%M:%S')
            difference = datetime_object - datetime_value  # difference between the present & the date on the excel sheet
            if (sheet.cell_value(count + 1, 2) == 'Implementation' or sheet.cell_value(count + 1, 2) == 'Assessment' or
                sheet.cell_value(count + 1, 2) == 'implementation' or sheet.cell_value(count + 1, 2) == 'assessment') \
                    and (sheet.cell_value(count + 1, 3) == 'Assigned' or sheet.cell_value(count + 1, 3) == 'assigned') \
                    and (difference.days >= 3):
                addrTo = [None]
                for counter in range(sheetAddr.nrows):
                    try:
                        if names == sheetAddr.cell_value(counter, 0):
                            addrTo = sheetAddr.cell_value(counter, 1)
                            break
                        if names == '':
                            print("Name not found!")
                            loopBreak = 1
                            break
                        if counter == sheetAddr.nrows - 1:
                            print("Name not found!2")
                            loopBreak = 1
                            break
                    except Exception:
                        print("Name not found!")

                if loopBreak == 1:
                    addrTo = 'RS@ontario.ca'

                addrFrom = 'DC@ontario.ca'  # Sender's email address
                message = MIMEMultipart()
                message['From'] = addrFrom
                message['To'] = addrTo
                message['Subject'] = "Incompletion of Task."
                if loopBreak == 1:
                    message['Subject'] = "Name Not Found! Incompletion of Task."
                body = """Hello """ + names + (", \n \nPlease complete your Implementation in ") + cid + (
                    ".\n \n"
                    "This ") + tid + (" has been active since ") + str(datetime_value) + (".\n \n"
                     "If there are issues with the change, please notify the requester and Operations Team lead.\n \n \n \n"
                     "Thank You,\n\n"
                     "Information Technology Services\n\n"
                     "Data Centre Networks\n\n"
                     "Senior Manager\n\n"
                    )

                message.attach(MIMEText(body, 'plain'))
                text = message.as_string()

                try:
                    smtpObj = smtplib.SMTP('relay')
                    smtpObj.sendmail(addrFrom, addrTo, text)
                    print("Successfully sent email")
                except smtplib.SMTPException:
                    print("Error: unable to send email")


def WriteInMasterSheet(sheet, sheet_write):
    # todo This function is responsible for updating the Master Spread Sheet and increments the counter based on
    # todo Task ID
    wb = load_workbook("MasterSpreadSheet.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    count = 2

    for rows in range(1, sheet.nrows):
        time_value = float(sheet.cell_value(rows, 6))  # Time is in float
        x = time_value
        datetime_value = datetime(*xlrd.xldate_as_tuple(x, 0))  # Converts to date time format
        timeNow = f"{datetime.now():%Y-%d-%m %H:%M:%S}"
        datetime_object = datetime.strptime(timeNow, '%Y-%d-%m %H:%M:%S')
        difference = datetime_object - datetime_value   # difference between the present & the date on the excel sheet

        if (sheet.cell_value(rows, 2) == 'Implementation' or sheet.cell_value(rows, 2) == 'Assessment' or
            sheet.cell_value(rows, 2) == 'implementation' or sheet.cell_value(rows, 2) == 'assessment') \
                and (sheet.cell_value(rows, 3) == 'Assigned' or sheet.cell_value(rows, 3) == 'assigned') \
                and (difference.days >= 3):
            count += 1
            for cols in range(0, sheet.ncols):
                if cols != sheet.ncols - 1:
                    Sheet1.cell(row=sheet_write.nrows + count, column=cols + 1).value = sheet.cell_value(rows, cols)
                if cols == sheet.ncols - 1:
                    time_value = float(sheet.cell_value(rows, 6))  # Time is in float
                    x = time_value
                    datetime_value = datetime(*xlrd.xldate_as_tuple(x, 0))  # Converts to date time format
                    Sheet1.cell(row=sheet_write.nrows + count, column=7).value = str(datetime_value)

                taskID = sheet.cell_value(rows, 1)

                counter = 0

                for i in range(0, sheet_write.nrows):
                    if sheet_write.cell_value(i, 1) == taskID:
                        counter += 1

                Sheet1.cell(row=sheet_write.nrows + count, column=8).value = counter + 1

    wb.save("MasterSpreadSheet.xlsx")


def attachmentToNick():
    # todo This function is responsible for sending the updated Master Spread Sheet to NK
    email_user = 'DCN@ontario.ca'
    email_send = 'RS@ontario.ca'

    msg = MIMEMultipart()
    msg['Subject'] = 'Task Reminder Master Spread Sheet'
    body = ""
    msg.attach(MIMEText(body, 'plain'))

    filename = 'MasterSpreadSheet.xlsx'
    attachment = open(filename, 'rb')

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= " + filename)

    msg.attach(part)
    text = msg.as_string()

    try:
        smtpObj = smtplib.SMTP('eesmgmtrelay.gov.on.ca')
        smtpObj.sendmail(email_user, email_send, text)
        print("Attachment sent!")
    except SMTPException:
        print("Error: unable to send attachment")


def main():
    # todo Calling functions in the required order
    sheet = locationTaskReminder()
    sheetAddr = locationDCNEmailAddress()
    sheet_write = locationMasterSpreadSheet_Read()

    Email_System(sheet, sheetAddr)

    WriteInMasterSheet(sheet, sheet_write)

    attachmentToNick()


main()
